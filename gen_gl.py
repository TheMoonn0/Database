import streamlit as st
import pandas as pd
import os
import re
import zipfile
import tempfile
import io
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# =========================
# Helper Functions
# =========================
def excel_col_to_index(col_str: str) -> int:
    num = 0
    for c in col_str:
        if c.isdigit():
            continue
        num = num * 26 + (ord(c.upper()) - ord("A")) + 1
    return num - 1

def convert_implied_decimal(val):
    """
    ‡πÅ‡∏õ‡∏•‡∏á implied decimal (‡∏´‡∏≤‡∏£ 100) ‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏Å‡∏£‡∏ì‡∏µ:
    - ‡πÄ‡∏õ‡πá‡∏ô‡πÄ‡∏•‡∏Ç‡∏•‡πâ‡∏ß‡∏ô ‡πÅ‡∏•‡∏∞
    - ‡∏°‡∏µ "00" ‡∏ô‡∏≥‡∏´‡∏ô‡πâ‡∏≤
    ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡πÄ‡∏Ç‡πâ‡∏≤‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç -> ‡∏Ñ‡∏∑‡∏ô‡∏Ñ‡πà‡∏≤‡∏ï‡∏£‡∏á ‡πÜ (‡πÑ‡∏°‡πà‡πÅ‡∏õ‡∏•‡∏á)
    """
    try:
        if val is None:
            return val
        val_str = str(val).strip()
        if (not val_str.isdigit()) or (not val_str.startswith("00")):
            return val_str
        return float(val_str) / 100.0
    except:
        return val

def extract_seq_num(val):
    text = str(val)
    match = re.search(r"seq_num:(\d+)", text)
    if match:
        return match.group(1)
    return str(val).strip()

def parse_dates_from_filename(filename: str):
    """
    Example:
      GLJV20251221_ATMI_SCB_000_ATMI1-D251110.csv
    - D-date: D251110 -> "251110"
    - JV-date: JV20251221 -> "251221" (YYMMDD)
    """
    base = os.path.basename(filename)

    d_match = re.search(r"[-_]?D(?P<d>\d{6})", base, flags=re.IGNORECASE)
    d_date = d_match.group("d") if d_match else None

    jv_match = re.search(r"JV(?P<jv>\d{8})", base, flags=re.IGNORECASE)
    jv_date = None
    if jv_match:
        jv_full = jv_match.group("jv")  # YYYYMMDD
        jv_date = jv_full[2:4] + jv_full[4:6] + jv_full[6:8]  # YYMMDD

    return d_date, jv_date

def pick_latest_files_by_duplicate_d_date(folder_path: str, files_list: list):
    """
    ‡∏ñ‡πâ‡∏≤ D-date ‡∏ã‡πâ‡∏≥‡∏Å‡∏±‡∏ô ‡πÉ‡∏´‡πâ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå‡∏ó‡∏µ‡πà JV-date (YYMMDD) ‡∏°‡∏≤‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
    ‡∏£‡∏≠‡∏á‡∏£‡∏±‡∏ö‡πÑ‡∏ü‡∏•‡πå: .csv, .trf, .txt, .xls, .xlsx
    """
    chosen = {}
    valid_exts = (".csv", ".trf", ".txt", ".xls", ".xlsx")

    for fn in files_list:
        file_path = os.path.join(folder_path, fn)
        if not fn.lower().endswith(valid_exts):
            continue
        if not os.path.isfile(file_path):
            continue

        d_date, jv_date = parse_dates_from_filename(fn)
        jv_int = int(jv_date) if jv_date and jv_date.isdigit() else -1

        if d_date is None:
            key = f"__NO_D__::{fn}"
            chosen[key] = {"file": file_path, "d_date": None, "jv_date": jv_date}
            continue

        if d_date not in chosen:
            chosen[d_date] = {"file": file_path, "d_date": d_date, "jv_date": jv_date, "_jv_int": jv_int}
        else:
            if jv_int > chosen[d_date].get("_jv_int", -1):
                chosen[d_date] = {"file": file_path, "d_date": d_date, "jv_date": jv_date, "_jv_int": jv_int}

    results = []
    for _, v in chosen.items():
        v.pop("_jv_int", None)
        results.append(v)

    results.sort(key=lambda x: os.path.basename(x["file"]).lower())
    return results

def strip_d_suffix_for_tlf_sheet(name_no_ext: str):
    return re.sub(r"[-_]?D\d{6}.*$", "", name_no_ext, flags=re.IGNORECASE).strip()

def make_unique_sheet_name(book, desired_name: str):
    base = (desired_name or "Sheet")[:31]
    name = base
    i = 2
    while name in book.sheetnames:
        suffix = f"_{i}"
        name = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    return name

def max_k_from_searchkey(series: pd.Series) -> int:
    """
    ‡∏´‡∏≤ max ‡πÄ‡∏•‡∏Ç‡∏ó‡πâ‡∏≤‡∏¢‡∏Ç‡∏≠‡∏á‡∏£‡∏π‡∏õ‡πÅ‡∏ö‡∏ö xxxx|k ‡∏à‡∏≤‡∏Å‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå _SearchKey
    ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡πÄ‡∏à‡∏≠‡∏Ñ‡∏∑‡∏ô 1
    """
    max_k = 1
    try:
        k_series = series.astype(str).str.extract(r"\|(\d+)\s*$")[0]
        k_series = pd.to_numeric(k_series, errors="coerce")
        max_k_val = k_series.max()
        if pd.notna(max_k_val):
            max_k = int(max_k_val)
    except:
        max_k = 1
    return max_k

# =========================
# Config
# =========================
tlf_reserved_rows = 2   # ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥ (‡πÅ‡∏ï‡πà‡∏à‡∏∞‡∏Ç‡∏¢‡∏≤‡∏¢‡∏ï‡∏≤‡∏° max "|k")
gl_reserved_rows = 10   # ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥ (‡πÅ‡∏ï‡πà‡∏à‡∏∞‡∏Ç‡∏¢‡∏≤‡∏¢‡∏ï‡∏≤‡∏° max "|k")
gap_rows = 3
exclude_tlf_columns = ["from_acct", "to_acct", "auth_branch_from"]

TLF_LABEL = "Database(ATMI)"

# GL columns (from file): J K L M N P AM AN AZ
gl_columns_letters = ["J", "K", "L", "M", "N", "P", "AM", "AN", "AZ"]

# ‚úÖ ‡πÇ‡∏Ñ‡∏£‡∏á‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå ‚Äú‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢‚Äù ‡∏Ç‡∏≠‡∏á‡∏ï‡∏≤‡∏£‡∏≤‡∏á ATMI (‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô GL_V4)
gl_final_headers = ["RC", "OC", "CH", "Product Code", "Account Code", "Tax", "DR", "CR", "Seq", "Details"]

# TLF columns
tlf_columns_letters = [
    "F", "G", "I", "J", "K", "M", "O", "V",
    "AF", "AS", "AT", "AU", "AV", "AX", "AZ", "CU", "DP", "BH"
]

gl_indices = [excel_col_to_index(c) for c in gl_columns_letters]
tlf_indices = [excel_col_to_index(c) for c in tlf_columns_letters]

def get_col_pos_in_tlf(target_letter):
    sorted_letters = sorted(tlf_columns_letters, key=lambda x: excel_col_to_index(x))
    try:
        return sorted_letters.index(target_letter)
    except:
        return -1

pos_AZ = get_col_pos_in_tlf("AZ")
pos_CU = get_col_pos_in_tlf("CU")

# Styles
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
header_font = Font(bold=True)
title_font = Font(bold=True, size=14, color="000000")
search_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# =========================
# Core Processing (In-Memory)
# =========================
def process_data_in_memory(db_path: str, source_files_list: list, temp_folder: str):
    output = io.BytesIO()

    try:
        with pd.ExcelFile(db_path) as db_book:
            files_to_process = pick_latest_files_by_duplicate_d_date(temp_folder, source_files_list)
            if not files_to_process:
                return None, "‡πÑ‡∏°‡πà‡∏û‡∏ö‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• (GL/TRF/CSV/TXT/Excel) ‡∏ó‡∏µ‡πà‡∏ñ‡∏π‡∏Å‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ô ZIP"

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for item in files_to_process:
                    file_path = item["file"]
                    filename = os.path.basename(file_path)
                    chosen_d_date = item["d_date"]

                    desired_sheet_name = chosen_d_date if chosen_d_date else os.path.splitext(filename)[0]

                    clean_name = re.sub(r"GL", "", filename, flags=re.IGNORECASE)
                    clean_name = os.path.splitext(clean_name)[0].strip()
                    fallback_lookup_name = strip_d_suffix_for_tlf_sheet(clean_name)

                    db_lookup_candidates = []
                    if chosen_d_date:
                        db_lookup_candidates.append(chosen_d_date)
                        db_lookup_candidates.append("D" + chosen_d_date)
                    db_lookup_candidates.append(fallback_lookup_name)

                    db_sheet_to_use = None
                    for cand in db_lookup_candidates:
                        if cand and cand in db_book.sheet_names:
                            db_sheet_to_use = cand
                            break

                    try:
                        # ---------- Load Database ----------
                        db_df = pd.DataFrame()
                        effective_db_reserved_rows = tlf_reserved_rows
                        max_k_db = 1

                        if db_sheet_to_use:
                            db_df = pd.read_excel(
                                db_book,
                                sheet_name=db_sheet_to_use,
                                usecols=tlf_indices,
                                dtype=str,
                            )
                            for col in db_df.columns:
                                db_df[col] = db_df[col].astype(str).str.strip()

                            # implied decimal (only when startswith "00")
                            if pos_AZ != -1 and pos_AZ < len(db_df.columns):
                                db_df.iloc[:, pos_AZ] = db_df.iloc[:, pos_AZ].apply(convert_implied_decimal)
                            if pos_CU != -1 and pos_CU < len(db_df.columns):
                                db_df.iloc[:, pos_CU] = db_df.iloc[:, pos_CU].apply(convert_implied_decimal)

                            # _SearchKey
                            if not db_df.empty and len(db_df.columns) > 8:
                                search_col = db_df.iloc[:, 8].astype(str).str.strip()
                                db_df["_SearchKey"] = search_col + "|" + (db_df.groupby(search_col).cumcount() + 1).astype(str)

                                # ‚úÖ ‡∏Ç‡∏¢‡∏≤‡∏¢ UI rows ‡∏ï‡∏≤‡∏° max|k (‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô GL_V4)
                                max_k_db = max_k_from_searchkey(db_df["_SearchKey"])
                                effective_db_reserved_rows = max(tlf_reserved_rows, max_k_db)

                        # ---------- Load Source (ATMI / GL) ----------
                        if filename.lower().endswith((".xls", ".xlsx")):
                            with pd.ExcelFile(file_path) as source_book:
                                raw_gl = pd.read_excel(source_book, header=None, usecols=gl_indices, dtype=str)
                        else:
                            try:
                                raw_gl = pd.read_csv(
                                    file_path,
                                    header=None,
                                    usecols=gl_indices,
                                    encoding="utf-8",
                                    dtype=str,
                                    engine="python",
                                )
                            except:
                                raw_gl = pd.read_csv(
                                    file_path,
                                    header=None,
                                    usecols=gl_indices,
                                    encoding="cp874",
                                    dtype=str,
                                    engine="python",
                                )

                        # ‚úÖ ‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô GL_V4: ‡∏ï‡∏±‡πâ‡∏á‡∏ä‡∏∑‡πà‡∏≠ source headers ‡πÅ‡∏•‡πâ‡∏ß‡∏™‡∏£‡πâ‡∏≤‡∏á Details/Seq ‡∏à‡∏≤‡∏Å AZ_RAW
                        gl_source_headers = ["RC", "OC", "CH", "Product Code", "Account Code", "Tax", "DR", "CR", "AZ_RAW"]
                        if len(raw_gl.columns) == len(gl_source_headers):
                            raw_gl.columns = gl_source_headers
                        else:
                            # ‡∏´‡∏≤‡∏Å‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÑ‡∏°‡πà‡∏ï‡∏£‡∏á ‡πÉ‡∏´‡πâ‡∏û‡∏¢‡∏≤‡∏¢‡∏≤‡∏°‡∏ï‡∏±‡πâ‡∏á‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏ó‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏ó‡∏≥‡πÑ‡∏î‡πâ
                            raw_gl.columns = gl_source_headers[: len(raw_gl.columns)]

                        gl_df = raw_gl.copy()
                        if "AZ_RAW" in gl_df.columns:
                            gl_df["Details"] = gl_df["AZ_RAW"]
                            gl_df["Seq"] = gl_df["AZ_RAW"].apply(extract_seq_num).astype(str).str.strip()
                        else:
                            gl_df["Details"] = ""
                            gl_df["Seq"] = ""

                        if "RC" in gl_df.columns:
                            gl_df["RC"] = gl_df["RC"].astype(str).str.strip()
                        if "CH" in gl_df.columns:
                            gl_df["CH"] = gl_df["CH"].astype(str).str.strip()

                        if "DR" in gl_df.columns:
                            gl_df["DR"] = pd.to_numeric(gl_df["DR"], errors="coerce").fillna(0)
                        if "CR" in gl_df.columns:
                            gl_df["CR"] = pd.to_numeric(gl_df["CR"], errors="coerce").fillna(0)

                        # ‚úÖ ‡∏Ñ‡∏á‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÅ‡∏ö‡∏ö GL_V4
                        for col in gl_final_headers:
                            if col not in gl_df.columns:
                                gl_df[col] = ""
                        gl_df = gl_df[gl_final_headers]

                        # Sort
                        cols_to_sort = ["CH", "RC", "OC", "Product Code"]
                        valid_sort_cols = [c for c in cols_to_sort if c in gl_df.columns]
                        if valid_sort_cols:
                            gl_df = gl_df.sort_values(by=valid_sort_cols, ascending=[True] * len(valid_sort_cols))

                        # _SearchKey (GL)
                        effective_gl_reserved_rows = gl_reserved_rows
                        max_k_gl = 1
                        if not gl_df.empty:
                            search_col_gl = gl_df["Seq"].astype(str)
                            gl_df["_SearchKey"] = search_col_gl + "|" + (gl_df.groupby(search_col_gl).cumcount() + 1).astype(str)

                            # ‚úÖ ‡∏Ç‡∏¢‡∏≤‡∏¢ UI rows ‡∏ï‡∏≤‡∏° max|k (‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô GL_V4)
                            max_k_gl = max_k_from_searchkey(gl_df["_SearchKey"])
                            effective_gl_reserved_rows = max(gl_reserved_rows, max_k_gl)

                        # ---------- Write Layout ----------
                        target_sheet_name = make_unique_sheet_name(writer.book, desired_sheet_name)
                        worksheet = writer.book.create_sheet(target_sheet_name)
                        writer.sheets[target_sheet_name] = worksheet
                        ws = writer.sheets[target_sheet_name]

                        search_ui_start_row = 1
                        db_ui_height = 2 + (effective_db_reserved_rows if not db_df.empty else 0)
                        gl_ui_height = 2 + (effective_gl_reserved_rows if not gl_df.empty else 0)
                        raw_data_start_row = search_ui_start_row + db_ui_height + gap_rows + gl_ui_height + 5

                        current_raw_row = raw_data_start_row

                        # ranges
                        db_data_start = db_data_end = None
                        db_key_col_letter = "A"
                        gl_data_start = gl_data_end = None
                        gl_key_col_letter = "A"

                        # --- Raw Database ---
                        if not db_df.empty:
                            ws.cell(row=current_raw_row - 1, column=1, value=TLF_LABEL).font = Font(bold=True, italic=True)
                            db_df.to_excel(writer, sheet_name=target_sheet_name, startrow=current_raw_row - 1, index=False)

                            db_data_start = current_raw_row + 1
                            db_data_end = current_raw_row + len(db_df)
                            db_key_col_letter = get_column_letter(len(db_df.columns))

                            for row in range(current_raw_row, db_data_end + 1):
                                for col in range(1, len(db_df.columns)):
                                    cell = ws.cell(row=row, column=col)
                                    cell.border = thin_border
                                    if row == current_raw_row:
                                        cell.alignment = align_center
                                        cell.font = header_font
                                    else:
                                        cell.alignment = align_right if isinstance(cell.value, (int, float)) else align_center
                                        if col == 9:
                                            cell.number_format = "@"
                            current_raw_row += len(db_df) + 4

                        # --- Raw ATMI ---
                        if not gl_df.empty:
                            ws.cell(row=current_raw_row - 1, column=1, value="--- Raw ATMI Data ---").font = Font(bold=True, italic=True)
                            gl_df.to_excel(writer, sheet_name=target_sheet_name, startrow=current_raw_row - 1, index=False)

                            gl_data_start = current_raw_row + 1
                            gl_data_end = current_raw_row + len(gl_df)
                            gl_key_col_letter = get_column_letter(len(gl_df.columns))  # includes _SearchKey too (because exported)

                            # Styling
                            for row in range(current_raw_row, gl_data_end + 1):
                                for col in range(1, len(gl_df.columns) + 1):
                                    cell = ws.cell(row=row, column=col)
                                    cell.border = thin_border
                                    if row == current_raw_row:
                                        cell.alignment = align_center
                                        cell.font = header_font
                                    else:
                                        col_name = gl_df.columns[col - 1]
                                        if col_name in ["DR", "CR"]:
                                            cell.alignment = align_right
                                            cell.number_format = "#,##0.00"
                                        elif col_name == "Details":
                                            cell.alignment = align_left
                                            cell.number_format = "@"
                                        else:
                                            cell.alignment = align_center
                                        if col_name in ["Seq", "Details"]:
                                            cell.number_format = "@"

                        # --- Search UI ---
                        ws[f"A{search_ui_start_row}"] = "üîç ‡∏Ñ‡πâ‡∏ô‡∏´‡∏≤‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• SEQ"
                        ws[f"A{search_ui_start_row}"].font = Font(bold=True, size=12)
                        ws[f"A{search_ui_start_row}"].alignment = Alignment(horizontal="right")

                        input_cell_ref = f"$B${search_ui_start_row}"
                        input_cell = ws[f"B{search_ui_start_row}"]
                        input_cell.fill = search_fill
                        input_cell.border = thin_border
                        input_cell.alignment = align_center
                        input_cell.number_format = "@"

                        report_row = search_ui_start_row + 2

                        # --- Database Report ---
                        if not db_df.empty:
                            ws[f"A{report_row}"] = TLF_LABEL
                            ws[f"A{report_row}"].font = title_font

                            display_cols = [c for c in db_df.columns if c != "_SearchKey" and c not in exclude_tlf_columns]

                            # swap (as original)
                            if "amt_1_full" in display_cols and "resp_byte" in display_cols:
                                idx1 = display_cols.index("amt_1_full")
                                idx2 = display_cols.index("resp_byte")
                                display_cols[idx1], display_cols[idx2] = display_cols[idx2], display_cols[idx1]

                            db_key_range_str = f"${db_key_col_letter}${db_data_start}:${db_key_col_letter}${db_data_end}"

                            for i, col_name in enumerate(display_cols, 1):
                                cell = ws.cell(row=report_row + 1, column=i)
                                cell.value = col_name
                                cell.font = Font(bold=True)
                                cell.border = thin_border
                                cell.alignment = align_center
                                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

                            data_start_row = report_row + 2

                            # ‚úÖ ‡πÉ‡∏ä‡πâ effective_db_reserved_rows (‡∏ï‡∏≤‡∏° max|k)
                            for r_offset in range(effective_db_reserved_rows):
                                current_formula_row = data_start_row + r_offset
                                k_value = r_offset + 1
                                match_logic = f'MATCH({input_cell_ref}&"|"&{k_value}, {db_key_range_str}, 0)'

                                for i, col_name in enumerate(display_cols, 1):
                                    original_col_idx = db_df.columns.get_loc(col_name)
                                    col_letter = get_column_letter(original_col_idx + 1)
                                    data_col_range = f"${col_letter}${db_data_start}:${col_letter}${db_data_end}"
                                    formula = f'=IFERROR(INDEX({data_col_range}, {match_logic}), "")'

                                    cell = ws.cell(row=current_formula_row, column=i)
                                    cell.value = formula
                                    cell.border = thin_border
                                    cell.alignment = align_center

                            report_row = data_start_row + effective_db_reserved_rows

                        report_row += gap_rows

                        # --- ATMI Report ---
                        if not gl_df.empty:
                            ws[f"A{report_row}"] = "ATMI"
                            ws[f"A{report_row}"].font = title_font

                            gl_display_cols = [c for c in gl_df.columns if c != "_SearchKey"]

                            for i, col_name in enumerate(gl_display_cols, 1):
                                cell = ws.cell(row=report_row + 1, column=i)
                                cell.value = col_name
                                cell.font = Font(bold=True)
                                cell.border = thin_border
                                cell.alignment = align_center
                                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

                            data_start_row = report_row + 2

                            # IMPORTANT: raw table includes _SearchKey col; key is last col
                            # For matching, we need range of _SearchKey column specifically.
                            # Find where _SearchKey was written in excel: it's the last column in exported gl_df.
                            gl_key_col_excel = get_column_letter(len(gl_df.columns))
                            gl_key_range_str = f"${gl_key_col_excel}${gl_data_start}:${gl_key_col_excel}${gl_data_end}"

                            # ‚úÖ ‡πÉ‡∏ä‡πâ effective_gl_reserved_rows (‡∏ï‡∏≤‡∏° max|k)
                            for r_offset in range(effective_gl_reserved_rows):
                                current_formula_row = data_start_row + r_offset
                                k_value = r_offset + 1
                                match_logic = f'MATCH({input_cell_ref}&"|"&{k_value}, {gl_key_range_str}, 0)'

                                for col_idx, col_name in enumerate(gl_display_cols, 1):
                                    col_letter = get_column_letter(col_idx)
                                    data_col_range = f"${col_letter}${gl_data_start}:${col_letter}${gl_data_end}"
                                    formula = f'=IFERROR(INDEX({data_col_range}, {match_logic}), "")'

                                    cell = ws.cell(row=current_formula_row, column=col_idx)
                                    cell.value = formula
                                    cell.border = thin_border

                                    if col_name in ["DR", "CR"]:
                                        cell.number_format = "#,##0.00"
                                        cell.alignment = align_right
                                    elif col_name == "Details":
                                        cell.number_format = "@"
                                        cell.alignment = align_left
                                    else:
                                        cell.alignment = align_center

                                    if col_name in ["Seq", "Details"]:
                                        cell.number_format = "@"

                        # --- Smart Auto Width (with Details locked) ---
                        col_widths = {}

                        def update_max_width(df, start_col_idx=1, skip_cols=None):
                            skip_cols = set(skip_cols or [])
                            for i, col_name in enumerate(df.columns):
                                if col_name in skip_cols:
                                    continue
                                current_idx = start_col_idx + i
                                max_len = len(str(col_name))
                                if not df.empty:
                                    try:
                                        data_len = df[col_name].astype(str).map(len).max()
                                        if pd.notna(data_len):
                                            max_len = max(max_len, data_len)
                                    except:
                                        pass
                                existing = col_widths.get(current_idx, 0)
                                col_widths[current_idx] = max(existing, max_len + 3)

                        if not db_df.empty:
                            update_max_width(db_df, start_col_idx=1)

                        if not gl_df.empty:
                            # skip Details to lock width later
                            update_max_width(gl_df, start_col_idx=1, skip_cols={"Details"})

                        for col_idx, width in col_widths.items():
                            col_letter = get_column_letter(col_idx)
                            final_width = max(12, min(width, 60))
                            writer.sheets[target_sheet_name].column_dimensions[col_letter].width = final_width

                        # widen A,B
                        writer.sheets[target_sheet_name].column_dimensions["A"].width = max(col_widths.get(1, 20), 30)
                        writer.sheets[target_sheet_name].column_dimensions["B"].width = max(col_widths.get(2, 20), 25)

                        # lock Details width
                        if "Details" in gl_df.columns:
                            details_col_idx = gl_df.columns.get_loc("Details") + 1
                            details_col_letter = get_column_letter(details_col_idx)
                            writer.sheets[target_sheet_name].column_dimensions[details_col_letter].width = 12

                    except Exception:
                        # continue next file
                        pass

                if "Sheet" in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
                    del writer.book["Sheet"]

    except Exception as e:
        return None, f"Error ‡∏≠‡πà‡∏≤‡∏ô‡πÑ‡∏ü‡∏•‡πå Database: {e}"

    output.seek(0)
    return output, None


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="Automated GL & Database Matching Report", layout="wide")
st.title("üìÇ Automated GL & Database Matching Report (Streamlit)")
st.write("‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå ZIP ‡∏ó‡∏µ‡πà‡∏õ‡∏£‡∏∞‡∏Å‡∏≠‡∏ö‡∏î‡πâ‡∏ß‡∏¢‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• (‡πÄ‡∏ä‡πà‡∏ô `.csv/.trf/.txt/.xls/.xlsx`) ‡πÅ‡∏•‡∏∞‡πÑ‡∏ü‡∏•‡πå `Database`")

uploaded_zip = st.file_uploader("Choose a ZIP file", type="zip")
output_name = st.text_input("Output Excel filename", value="GL_File.xlsx")

if uploaded_zip:
    if st.button("üöÄ Process Files"):
        with st.spinner("Extracting & Processing..."):
            with tempfile.TemporaryDirectory() as temp_dir:
                try:
                    # Extract ZIP
                    with zipfile.ZipFile(uploaded_zip, "r") as zip_ref:
                        zip_ref.extractall(temp_dir)

                    # Identify files
                    db_path = None
                    source_files = []

                    for root, _, files in os.walk(temp_dir):
                        for file in files:
                            if file.startswith(".") or "__MACOSX" in root:
                                continue

                            full_path = os.path.join(root, file)

                            # ‚úÖ ‡∏à‡∏±‡∏ö‡πÑ‡∏ü‡∏•‡πå‡∏ä‡∏∑‡πà‡∏≠ Database
                            if "DATABASE" in file.upper():
                                if db_path is None:
                                    db_path = full_path
                            else:
                                source_files.append(os.path.relpath(full_path, temp_dir))

                    if not db_path:
                        st.error("‚ùå ‡πÑ‡∏°‡πà‡∏û‡∏ö‡πÑ‡∏ü‡∏•‡πå Database ‡πÉ‡∏ô ZIP (‡∏ï‡πâ‡∏≠‡∏á‡∏°‡∏µ‡∏Ñ‡∏≥‡∏ß‡πà‡∏≤ 'Database' ‡πÉ‡∏ô‡∏ä‡∏∑‡πà‡∏≠‡πÑ‡∏ü‡∏•‡πå)")
                    elif not source_files:
                        st.error("‚ùå ‡πÑ‡∏°‡πà‡∏û‡∏ö‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• Source ‡πÉ‡∏ô ZIP")
                    else:
                        st.info(f"üìç Found Database: {os.path.basename(db_path)}")
                        st.info(f"üìç Found Source Files: {len(source_files)} files")

                        excel_file, error_msg = process_data_in_memory(db_path, source_files, temp_dir)

                        if error_msg:
                            st.error(error_msg)
                        else:
                            st.success("‚úÖ Processing Complete!")
                            st.download_button(
                                label="üì• Download Final Excel",
                                data=excel_file,
                                file_name=output_name.strip() or "GL_File.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )

                except zipfile.BadZipFile:
                    st.error("‚ùå ZIP ‡πÑ‡∏ü‡∏•‡πå‡πÄ‡∏™‡∏µ‡∏¢‡∏´‡∏£‡∏∑‡∏≠‡πÄ‡∏õ‡∏¥‡∏î‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ")
                except Exception as e:
                    st.error(f"‚ùå Error during processing: {e}")
